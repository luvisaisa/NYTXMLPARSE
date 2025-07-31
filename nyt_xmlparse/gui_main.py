"""Simple GUI application for parsing XML files and exporting results."""

from __future__ import annotations

import os
import tkinter as tk
from tkinter import filedialog, messagebox

from .parser import NYTXMLParser


class NYTXMLGuiApp:
    """GUI application to parse XML files and optionally export to Excel."""

    def __init__(self, master: tk.Tk) -> None:
        self.master = master
        self.master.title("NYT XML Parser")
        self.parser = NYTXMLParser()
        self.files: list[str] = []
        self._create_widgets()

    def _create_widgets(self) -> None:
        frame = tk.Frame(self.master)
        frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        btn_files = tk.Button(frame, text="Select XML Files", command=self.select_files)
        btn_files.pack(fill=tk.X)

        btn_folder = tk.Button(frame, text="Select Folder", command=self.select_folder)
        btn_folder.pack(fill=tk.X, pady=(5, 0))

        self.listbox = tk.Listbox(frame, height=6)
        self.listbox.pack(fill=tk.BOTH, expand=True, pady=5)

        parse_btn = tk.Button(frame, text="Parse", command=self.parse_files)
        parse_btn.pack(fill=tk.X)

    def select_files(self) -> None:
        filenames = filedialog.askopenfilenames(title="Choose XML files", filetypes=[("XML Files", "*.xml")])
        if filenames:
            self.files = list(filenames)
            self._update_file_list()

    def select_folder(self) -> None:
        folder = filedialog.askdirectory(title="Choose folder with XML files")
        if folder:
            self.files = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(".xml")]
            self._update_file_list()

    def _update_file_list(self) -> None:
        self.listbox.delete(0, tk.END)
        for path in self.files:
            self.listbox.insert(tk.END, os.path.basename(path))

    def parse_files(self) -> None:
        if not self.files:
            messagebox.showinfo("No files", "Please select XML files to parse.")
            return
        parsed = []
        for path in self.files:
            try:
                with open(path, "r", encoding="utf-8") as fh:
                    content = fh.read()
                root = self.parser.parse(content)
                parsed.append((os.path.basename(path), root.tag))
            except Exception as exc:  # pragma: no cover - GUI message
                messagebox.showerror("Parse error", f"Failed to parse {path}: {exc}")
                return
        self._ask_export(parsed)

    def _ask_export(self, data: list[tuple[str, str]]) -> None:
        if messagebox.askyesno("Export", "Export parsed data to Excel?"):
            self._export_to_excel(data)

    def _export_to_excel(self, data: list[tuple[str, str]]) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            messagebox.showerror("Missing dependency", "openpyxl is required to export to Excel.")
            return

        path = filedialog.asksaveasfilename(
            title="Save Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not path:
            return

        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Filename", "RootTag"])

        for filename, tag in data:
            ws.append([filename, tag])

        wb.save(path)
        messagebox.showinfo("Exported", f"Data exported to {path}")


def main() -> None:
    root = tk.Tk()
    NYTXMLGuiApp(root)
    root.mainloop()


if __name__ == "__main__":  # pragma: no cover - manual execution
    main()
